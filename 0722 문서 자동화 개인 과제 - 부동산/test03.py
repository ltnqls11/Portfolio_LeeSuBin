# 1️⃣ 라이브러리 설치
# pip install pandas openpyxl

# 2️⃣ 엑셀 파일 준비
# 내_보고서_양식.xlsx
# 예: A1 셀부터 표가 시작되어 있어야 함. (헤더 포함)

# 3️⃣ 코드 예시
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
from datetime import datetime
import glob

def find_csv_file():
    """최신 CSV 파일을 찾는 함수"""
    csv_files = glob.glob('수원시_*_매매_전세_평균가_*.csv')
    if not csv_files:
        csv_files = glob.glob('수원시_1년_매매_전세_평균가.csv')
    
    if not csv_files:
        raise FileNotFoundError("CSV 파일을 찾을 수 없습니다. test02.py를 먼저 실행하세요.")
    
    # 가장 최신 파일 선택
    latest_file = max(csv_files, key=os.path.getctime)
    print(f"사용할 CSV 파일: {latest_file}")
    return latest_file

def create_excel_template(file_path):
    """엑셀 템플릿 파일을 생성하는 함수"""
    wb = Workbook()
    ws = wb.active
    ws.title = "수원시_부동산_데이터"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 추가
    headers = ["구", "월", "매매 평균 (억원)", "전세 평균 (억원)", "매매 건수", "전세 건수"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 열 너비 조정
    column_widths = [10, 12, 18, 18, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    wb.save(file_path)
    print(f"엑셀 템플릿 파일이 생성되었습니다: {file_path}")
    return wb

def update_excel_with_data(df, file_path, sheet_name="수원시_부동산_데이터"):
    """엑셀 파일에 데이터를 업데이트하는 함수"""
    try:
        # 기존 엑셀 파일 열기
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            print(f"기존 엑셀 파일을 열었습니다: {file_path}")
        else:
            print(f"엑셀 파일이 없어서 새로 생성합니다: {file_path}")
            wb = create_excel_template(file_path)
        
        # 시트 확인 및 생성
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"기존 시트 '{sheet_name}'를 사용합니다.")
        else:
            ws = wb.create_sheet(sheet_name)
            print(f"새 시트 '{sheet_name}'를 생성했습니다.")
        
        # 기존 데이터 삭제 (헤더 제외)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
            print("기존 데이터를 삭제했습니다.")
        
        # 헤더가 없으면 추가
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            headers = ["구", "월", "매매 평균 (억원)", "전세 평균 (억원)", "매매 건수", "전세 건수"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # 데이터 추가 (헤더 제외)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 숫자 데이터 가운데 정렬
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="center")
        
        # 테두리 추가
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
        
        # 저장
        wb.save(file_path)
        print(f"엑셀 파일이 성공적으로 업데이트되었습니다: {file_path}")
        print(f"총 {len(df)}행의 데이터가 추가되었습니다.")
        
        return True
        
    except Exception as e:
        print(f"엑셀 업데이트 중 오류 발생: {e}")
        return False

def main():
    """메인 실행 함수"""
    try:
        print("=== 엑셀 보고서 생성 시작 ===")
        
        # 1. CSV 파일 찾기 및 읽기
        csv_file = find_csv_file()
        df = pd.read_csv(csv_file, encoding='utf-8-sig')
        print(f"CSV 데이터 로드 완료: {len(df)}행")
        
        # 2. 엑셀 파일 경로 설정
        today = datetime.now().strftime("%Y%m%d")
        file_path = f'수원시_부동산_보고서_{today}.xlsx'
        
        # 3. 엑셀 업데이트
        success = update_excel_with_data(df, file_path)
        
        if success:
            print("\n=== 보고서 생성 완료 ===")
            print(f"파일 위치: {os.path.abspath(file_path)}")
            
            # 간단한 통계 출력
            print(f"\n=== 데이터 요약 ===")
            print(f"총 데이터 수: {len(df)}행")
            print(f"수집 기간: {df['월'].min()} ~ {df['월'].max()}")
            print(f"대상 지역: {', '.join(df['구'].unique())}")
        else:
            print("보고서 생성에 실패했습니다.")
            
    except FileNotFoundError as e:
        print(f"파일 오류: {e}")
        print("해결 방법: test02.py를 먼저 실행하여 CSV 파일을 생성하세요.")
    except Exception as e:
        print(f"예상치 못한 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()