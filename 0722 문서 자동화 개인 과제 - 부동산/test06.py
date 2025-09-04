# 1️⃣ 필요한 라이브러리 설치
# pip install pandas openpyxl matplotlib seaborn

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import glob
import warnings
warnings.filterwarnings('ignore')

# 한글 폰트 설정
def setup_korean_font():
    """한글 폰트 설정 함수"""
    plt.rcParams.update({
        'font.family': ['Malgun Gothic', 'Microsoft YaHei', 'SimHei', 'sans-serif'],
        'axes.unicode_minus': False,
        'font.size': 11,
        'figure.titlesize': 14,
        'axes.titlesize': 12,
        'axes.labelsize': 11,
        'legend.fontsize': 10,
        'xtick.labelsize': 9,
        'ytick.labelsize': 9
    })
    print("✅ 한글 폰트 설정 완료")

# 한글 폰트 설정 실행
setup_korean_font()

def find_csv_file():
    """최신 CSV 파일을 찾는 함수"""
    csv_files = glob.glob('수원시_*_매매_전세_평균가_*.csv')
    if not csv_files:
        csv_files = glob.glob('수원시_1년_매매_전세_평균가.csv')
    
    if not csv_files:
        raise FileNotFoundError("CSV 파일을 찾을 수 없습니다. test02.py 또는 test02_demo.py를 먼저 실행하세요.")
    
    # 가장 최신 파일 선택
    latest_file = max(csv_files, key=os.path.getctime)
    print(f"사용할 CSV 파일: {latest_file}")
    return latest_file

def load_and_prepare_data(file_path):
    """데이터 로드 및 전처리"""
    try:
        # CSV 불러오기
        df = pd.read_csv(file_path, encoding='utf-8-sig')
        print(f"데이터 로드 완료: {len(df)}행")
        
        # 컬럼명 정리
        column_mapping = {
            '구': '구명',
            '월': '날짜',
            '매매 평균 (억원)': '매매 평균',
            '전세 평균 (억원)': '전세 평균'
        }
        df = df.rename(columns=column_mapping)
        
        # 필요한 컬럼 확인
        if '구명' not in df.columns and '구' in df.columns:
            df['구명'] = df['구']
        if '날짜' not in df.columns and '월' in df.columns:
            df['날짜'] = df['월']
        
        # 숫자 컬럼 변환
        df['매매 평균'] = pd.to_numeric(df['매매 평균'], errors='coerce')
        df['전세 평균'] = pd.to_numeric(df['전세 평균'], errors='coerce')
        
        # 날짜 형식 처리
        df['날짜'] = df['날짜'].astype(str)
        
        # 날짜를 YYYY-MM 형식으로 변환
        def format_date(date_str):
            try:
                if len(date_str) == 6 and date_str.isdigit():
                    year = date_str[:4]
                    month = date_str[4:6]
                    return f"{year}-{month}"
                elif '-' in date_str and len(date_str) == 7:
                    return date_str
                else:
                    return date_str
            except:
                return date_str
        
        df['날짜_표시'] = df['날짜'].apply(format_date)
        df = df.sort_values(['구명', '날짜'])
        
        print(f"구별 데이터 수: {df['구명'].value_counts().to_dict()}")
        print(f"분석 기간: {df['날짜_표시'].min()} ~ {df['날짜_표시'].max()}")
        
        return df
        
    except Exception as e:
        print(f"데이터 로드 오류: {e}")
        raise

def create_charts_for_excel(df):
    """엑셀용 차트 생성"""
    chart_files = []
    gu_list = df['구명'].unique()
    
    print(f"\n=== 엑셀용 차트 생성 ({len(gu_list)}개) ===")
    
    for gu in gu_list:
        try:
            data = df[df['구명'] == gu].copy()
            
            if data.empty:
                print(f"{gu}: 데이터 없음")
                continue
            
            # 날짜 정렬
            data = data.sort_values('날짜').reset_index(drop=True)
            
            # 차트 생성
            plt.figure(figsize=(12, 6))
            
            # 데이터 유효성 검사
            valid_trade = data['매매 평균'].dropna()
            valid_rent = data['전세 평균'].dropna()
            
            if not valid_trade.empty:
                plt.plot(data['날짜_표시'], data['매매 평균'], 
                        marker='o', linewidth=2.5, markersize=6, 
                        label='매매 평균', color='#2E86AB', alpha=0.8)
            
            if not valid_rent.empty:
                plt.plot(data['날짜_표시'], data['전세 평균'], 
                        marker='s', linewidth=2.5, markersize=6, 
                        label='전세 평균', color='#A23B72', alpha=0.8)
            
            plt.title(f'{gu} 최근 1년 매매/전세 평균가 추이', fontsize=14, fontweight='bold', pad=20)
            plt.xlabel('연월', fontsize=12)
            plt.ylabel('금액 (억원)', fontsize=12)
            plt.legend(fontsize=11, frameon=True, fancybox=True, shadow=True)
            plt.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            
            # 파일 저장
            img_file = f'{gu}_chart_excel.png'
            plt.savefig(img_file, dpi=200, bbox_inches='tight', facecolor='white')
            plt.close()
            
            chart_files.append(img_file)
            print(f"✅ {gu} 차트 생성: {img_file}")
            
        except Exception as e:
            print(f"❌ {gu} 차트 생성 오류: {e}")
            continue
    
    return chart_files

def create_excel_report(df, chart_files):
    """엑셀 보고서 생성"""
    try:
        print(f"\n=== 엑셀 보고서 생성 ===")
        
        # 워크북 생성
        wb = Workbook()
        
        # 첫 번째 시트: 데이터 요약
        ws_summary = wb.active
        ws_summary.title = '데이터_요약'
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 요약 정보 추가
        summary_data = [
            ['수원시 부동산 분석 보고서'],
            [''],
            ['생성일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['분석 기간', f"{df['날짜_표시'].min()} ~ {df['날짜_표시'].max()}"],
            ['분석 지역', ', '.join(df['구명'].unique())],
            ['총 데이터 수', len(df)],
            [''],
            ['구별 데이터 현황']
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 제목
                    cell.font = Font(bold=True, size=16)
                elif row_idx == 8:  # 소제목
                    cell.font = Font(bold=True, size=12)
        
        # 구별 통계 추가
        start_row = len(summary_data) + 2
        stats_headers = ['구', '매매 평균가', '전세 평균가', '데이터 수']
        
        for col_idx, header in enumerate(stats_headers, 1):
            cell = ws_summary.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 구별 통계 데이터
        for row_idx, gu in enumerate(df['구명'].unique(), start_row + 1):
            gu_data = df[df['구명'] == gu]
            trade_avg = gu_data['매매 평균'].mean()
            rent_avg = gu_data['전세 평균'].mean()
            data_count = len(gu_data)
            
            ws_summary.cell(row=row_idx, column=1, value=gu)
            ws_summary.cell(row=row_idx, column=2, value=f"{trade_avg:.2f}억원" if not pd.isna(trade_avg) else "N/A")
            ws_summary.cell(row=row_idx, column=3, value=f"{rent_avg:.2f}억원" if not pd.isna(rent_avg) else "N/A")
            ws_summary.cell(row=row_idx, column=4, value=data_count)
        
        # 두 번째 시트: 전체 데이터
        ws_data = wb.create_sheet(title='전체_데이터')
        
        # 데이터 추가
        headers = ['구명', '날짜', '매매 평균 (억원)', '전세 평균 (억원)']
        if '매매 건수' in df.columns:
            headers.extend(['매매 건수', '전세 건수'])
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 행 추가
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws_data.cell(row=row_idx, column=1, value=row['구명'])
            ws_data.cell(row=row_idx, column=2, value=row['날짜_표시'])
            ws_data.cell(row=row_idx, column=3, value=row['매매 평균'])
            ws_data.cell(row=row_idx, column=4, value=row['전세 평균'])
            
            if '매매 건수' in df.columns:
                ws_data.cell(row=row_idx, column=5, value=row.get('매매 건수', 0))
                ws_data.cell(row=row_idx, column=6, value=row.get('전세 건수', 0))
        
        # 세 번째 시트: 차트 모음
        ws_charts = wb.create_sheet(title='차트_모음')
        
        # 차트 제목
        ws_charts.cell(row=1, column=1, value='구별 부동산 가격 추이 차트').font = Font(bold=True, size=16)
        
        # 차트 이미지 삽입
        current_row = 3
        for chart_file in chart_files:
            if os.path.exists(chart_file):
                try:
                    img = Image(chart_file)
                    # 이미지 크기 조정
                    img.width = 600
                    img.height = 300
                    ws_charts.add_image(img, f'A{current_row}')
                    current_row += 20  # 다음 차트를 위한 공간
                    print(f"✅ 차트 삽입 완료: {chart_file}")
                except Exception as e:
                    print(f"❌ 차트 삽입 오류 ({chart_file}): {e}")
        
        # 열 너비 조정
        for ws in [ws_summary, ws_data]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        today = datetime.now().strftime("%Y%m%d")
        excel_filename = f'수원시_부동산_종합보고서_{today}.xlsx'
        wb.save(excel_filename)
        
        print(f"✅ 엑셀 보고서 생성 완료: {excel_filename}")
        return excel_filename
        
    except Exception as e:
        print(f"❌ 엑셀 보고서 생성 오류: {e}")
        raise

def cleanup_temp_files(chart_files):
    """임시 차트 파일 정리"""
    for chart_file in chart_files:
        try:
            if os.path.exists(chart_file):
                os.remove(chart_file)
                print(f"🗑️ 임시 파일 삭제: {chart_file}")
        except Exception as e:
            print(f"⚠️ 파일 삭제 오류 ({chart_file}): {e}")

def main():
    """메인 실행 함수"""
    try:
        print("=== 수원시 부동산 종합 보고서 생성 시작 ===")
        
        # 1. 데이터 로드
        csv_file = find_csv_file()
        df = load_and_prepare_data(csv_file)
        
        # 2. 차트 생성
        chart_files = create_charts_for_excel(df)
        
        # 3. 엑셀 보고서 생성
        excel_file = create_excel_report(df, chart_files)
        
        # 4. 임시 파일 정리
        cleanup_temp_files(chart_files)
        
        # 5. 결과 요약
        print(f"\n✅ 종합 보고서 생성 완료!")
        print(f"📁 생성된 파일: {excel_file}")
        print(f"📊 포함된 내용:")
        print(f"  - 데이터 요약 시트")
        print(f"  - 전체 데이터 시트")
        print(f"  - 구별 차트 모음 시트")
        print(f"📈 분석 기간: {df['날짜_표시'].min()} ~ {df['날짜_표시'].max()}")
        print(f"🏘️ 분석 지역: {', '.join(df['구명'].unique())}")
        
        return excel_file
        
    except FileNotFoundError as e:
        print(f"파일 오류: {e}")
        print("해결 방법: test02.py 또는 test02_demo.py를 먼저 실행하세요.")
    except Exception as e:
        print(f"예상치 못한 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    result = main()